import configparser
from scapy.all import *
import xlsxwriter
import os
import argparse
import csv
import openpyxl
import ast
from OTXv2 import OTXv2
import IndicatorTypes
import requests

def get_otxverdict(ip):
    otx=OTXv2(otx_apikey,server=OTX_SERVER)

    try:
        result = otx.get_indicator_details_by_section(IndicatorTypes.IPv4, ip, 'general')
        count = result['pulse_info']['count']
        return count
    except :
        return 0

def get_vtverdict(ip):
     try:
        vt_response = requests.request("GET", vt_url + str(ip), headers=vt_headers)
     except Exception as e:
         print(e)
     if (vt_response.status_code==429):
         print("\n")
         print("Virus Total API Quota exceeded.\n")
         exit(0)
     elif (vt_response.status_code ==200):
         vt_response = vt_response.json()
         try:
             vt_response = vt_response['data']['attributes']['last_analysis_stats']
             malicious=vt_response['malicious']
             suspicious= vt_response['suspicious']
             vt_result=[malicious,suspicious]
             return [malicious,suspicious]

         except Exception as e:
             return [0,0]

def data(args):
    outfile = openpyxl.load_workbook(args.output_file)
    sheet = outfile.active
    max_rows = sheet.max_row

    for row in range(1,max_rows):
        ip = sheet.cell(row + 1, 1).value
        vt_result=get_vtverdict(ip)
        otx_result=get_otxverdict(ip)
        sheet.cell(row+1,2).value=vt_result[0]
        sheet.cell(row+1,3).value=vt_result[1]
        sheet.cell(row+1,4).value=otx_result
    outfile.save(args.output_file)
    outfile.close()

def write_ips(args,iplist):
    outfile=openpyxl.load_workbook(args.output_file)
    sheet=outfile.active
    row=2
    for ip in iplist:
        if (re.match("(?!(10)|192\.168|172\.(2[0-9]|1[6-9]|3[0-1])|(25[6-9]|2[6-9][0-9]|[3-9][0-9][0-9]|99[1-9]))[0-9]{1,3}\.(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)",ip)):
            sheet.cell(row,1).value=ip
            row+=1
    outfile.save(args.output_file)
    outfile.close()


def read_pcap(args):
    iplist=[]
    packets = rdpcap(args.input_file)
    for packet_number in range(len(packets)):
            packet = packets[packet_number]
            try:
                srcip=packet['IP'].src
                destip=packet['IP'].dst
                iplist.append(srcip)
                iplist.append(destip)
            except Exception as e:
                print(f"Error reading IP at {packet_number} packet. \n")

    iplist=list(dict.fromkeys(iplist))
    write_ips(args,iplist)
    data(args)

def read_csv(args):
    iplist = []
    input = open(args.input_file)
    csvfile=csv.reader(input)
    for ip in csvfile:
        #csv values are retured as list, so below line will convert list to str.
        strip = ''.join(ip)
        iplist.append(strip)
    write_ips(args, iplist)
    data(args)

def read_txt(args):
    iplist=[]
    try:
        input=open(args.input_file,'r')
        for ips in input.readlines():
            iplist.append(ips.rstrip())
        write_ips(args,iplist)
        data(args)

    except Exception as e:
        print(e)

def get_filetype(args):
    if (os.path.splitext(args.input_file))[1]=='.txt':
        read_txt(args)

    elif(os.path.splitext(args.input_file))[1]=='.csv':
        read_csv(args)

    elif (os.path.splitext(args.input_file))[1]=='.pcap':
        read_pcap(args)
    else:
        print("\n")
        print("Wrong type of input file Provided. Only accepts ['txt','csv','pcap'].")


def create_file(args):
    if ((os.path.splitext(args.output_file))[1] !='.xlsx'):
        print("\n")
        print("Wrong output extension selected. Only supports 'xlsx' file type for output. ")
        exit(0)
    c_col=0
    columns=['IP','VT Malicious Count','VT Suspicious Count','OTX Pulse Count']
    workbook=xlsxwriter.Workbook(args.output_file,{'constant_memory':True})
    bold = workbook.add_format({'bold': True})
    worksheet=workbook.add_worksheet()
    for heading in columns:
        worksheet.write(0,c_col,heading,bold)
        c_col+=1
    workbook.close()
    get_filetype(args)

def check_arguments(args):
    if not (os.path.isfile(args.input_file)):
        print("\n")
        print("You need to provide correct input file.")
        exit()
    if (os.path.splitext(args.output_file))[1] !='.xlsx':
        print("only accept xlsx file as output.")
        exit()
    else:
        create_file(args)

def  get_arguments():

    parser = argparse.ArgumentParser()
    parser.add_argument('-f', '--file', dest='input_file', action="store")
    parser.add_argument('-o', '--output', dest='output_file', action='store')
    args = parser.parse_args()
    return args

if __name__ == '__main__':

    ''' 
    Description : Porus can parse pcap file and get IPs reputation from VT and OTX. You need to provide VT and OTX api key in config.ini file.
    config.ini file need to be in the same directory as porus.py.
    
    Author: Majid Jahangeer
    Email: mianmajid432@gmail.com
    Version: 1.1
    Usage : python porus.py -i sample.pcap -o sample.xlsx
            python porus.py -i ips.txt -o sample.xlsx
            python porus.py -i ips.csv -o results.xlsx
    '''
    vt_url = "https://www.virustotal.com/api/v3/ip_addresses/"
    OTX_SERVER = 'https://otx.alienvault.com/'
    config=configparser.ConfigParser()
    script_path=os.path.dirname(os.path.realpath(__file__))
    config.read(script_path+'\config.ini')
    vt_headers = config['VirusTotal']['vt_headers']
    vt_headers=ast.literal_eval(vt_headers)
    otx_apikey=config['OTX']['OTX_API_KEY']
    test=requests.request('GET',vt_url+'8.8.8.8',headers=vt_headers)
    if (test.status_code ==401):
        print("Virus Total Api Key incorrect.\n")
        exit(0)
    args=get_arguments()
    check_arguments(args)